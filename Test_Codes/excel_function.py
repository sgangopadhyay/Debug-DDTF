from openpyxl import load_workbook

class Suman_Excel_Function:
    def __init__(self, file_name):
        self.file = file_name
        self.wb = load_workbook(file_name)
        self.worksheet = self.wb.active
        self.row_count = self.worksheet.max_row
        self.column_count = self.worksheet.max_column    
    
    def dic(self):
        result = {}
        for row in range(2, self.worksheet.max_row+1):
            key = self.worksheet.cell(row, 1).value
            value = self.worksheet.cell(row, 2).value
            result[key]=value
        return(result)